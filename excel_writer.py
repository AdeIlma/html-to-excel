"""
excel_writer.py
----------------
Menulis DataFrame hasil parsing GoFood ke file Excel (.xlsx) yang rapi:
- Sheet "Semua Toko": gabungan semua produk dari semua toko
- Satu sheet tambahan per toko (kalau lebih dari 1 toko diupload)
- Header tebal, kolom harga format Rupiah, lebar kolom menyesuaikan isi
"""

import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
BODY_FONT = Font(name="Arial")
CURRENCY_FORMAT = '"Rp"#,##0'


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str):
    df_sorted = df.sort_values(["Kategori", "Produk"]).reset_index(drop=True)
    df_sorted.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    ws = writer.sheets[sheet_name]

    # Header styling
    for col_idx, col_name in enumerate(df_sorted.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Body font + currency format on "Harga" column
    harga_col_idx = list(df_sorted.columns).index("Harga") + 1
    for row_idx in range(2, len(df_sorted) + 2):
        for col_idx in range(1, len(df_sorted.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            if col_idx == harga_col_idx:
                cell.number_format = CURRENCY_FORMAT

    # Auto-width kolom sederhana berdasarkan panjang teks terpanjang
    for col_idx, col_name in enumerate(df_sorted.columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df_sorted.iloc[:, col_idx - 1]]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Ubah DataFrame gabungan (kolom: Toko, Produk, Kategori, Harga) menjadi
    file Excel dalam bentuk bytes, siap di-download lewat Streamlit.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet gabungan semua toko
        _write_sheet(writer, df, "Semua Toko")

        # Kalau lebih dari 1 toko, tambahkan sheet per toko juga
        toko_list = df["Toko"].unique().tolist()
        if len(toko_list) > 1:
            for toko in toko_list:
                sub_df = df[df["Toko"] == toko]
                # Nama sheet Excel maksimal 31 karakter & tidak boleh karakter tertentu
                safe_name = "".join(c for c in toko if c not in r'[]:*?/\\')[:31]
                _write_sheet(writer, sub_df, safe_name or "Toko")

    return output.getvalue()
